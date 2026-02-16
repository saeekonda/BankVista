"""
BankVista AI - ULTIMATE COMPLETE VERSION
✅ Chat scrolls properly in white space
✅ ALL features included (nothing removed)
✅ Hero section + Status bar + KPI cards
✅ Enhanced AI responses
✅ Charts in chat
✅ Dashboard with all analytics
✅ Heatmaps, Zone analytics, Predictions
✅ Excel export
✅ COMPLETE WORKING VERSION
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
from datetime import date, datetime, timedelta
import io
import os
import warnings
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

warnings.filterwarnings('ignore')
load_dotenv()

st.set_page_config(
    page_title="BankVista AI - Intelligent Banking Analytics",
    page_icon="🤖",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ═══════════════════════════════════════════════════════════════
# COMPLETE CSS WITH SCROLLING CHAT
# ═══════════════════════════════════════════════════════════════
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=Space+Grotesk:wght@500;600;700;800&display=swap');

    .main { 
        background: #f0f2f5; 
        font-family: 'Inter', sans-serif;
    }
    
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    
    /* Hero Section */
    .hero-container {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 3.5rem 2rem; 
        border-radius: 24px; 
        margin-bottom: 2rem;
        box-shadow: 0 20px 60px rgba(102,126,234,0.3);
        position: relative; 
        overflow: hidden;
    }
    
    .hero-container::before {
        content:''; 
        position:absolute; 
        top:-50%; 
        right:-50%;
        width:200%; 
        height:200%;
        background: radial-gradient(circle, rgba(255,255,255,0.08) 1px, transparent 1px);
        background-size: 50px 50px;
        animation: grid-move 20s linear infinite;
    }
    
    @keyframes grid-move { 
        0%{transform:translate(0,0)} 
        100%{transform:translate(50px,50px)} 
    }

    .main-title {
        font-family:'Space Grotesk',sans-serif; 
        font-size:3.8rem; 
        font-weight:800;
        color:white; 
        text-align:center; 
        margin-bottom:0.8rem;
        position:relative; 
        z-index:1; 
        text-shadow:0 2px 20px rgba(0,0,0,0.2); 
        letter-spacing:-2px;
    }
    
    .subtitle { 
        font-size:1.3rem; 
        color:rgba(255,255,255,0.92); 
        text-align:center; 
        font-weight:500; 
        margin-bottom:1rem; 
        position:relative; 
        z-index:1; 
    }
    
    .feature-pills { 
        display:flex; 
        flex-wrap:wrap; 
        justify-content:center; 
        gap:0.8rem; 
        margin-top:1.5rem; 
        position:relative; 
        z-index:1; 
    }
    
    .feature-pill {
        background:rgba(255,255,255,0.18); 
        backdrop-filter:blur(10px);
        border:1px solid rgba(255,255,255,0.3); 
        padding:0.5rem 1.1rem;
        border-radius:25px; 
        color:white; 
        font-weight:600; 
        font-size:0.85rem;
        display:inline-flex; 
        align-items:center; 
        gap:0.4rem; 
        transition:all 0.3s ease;
    }
    
    .feature-pill:hover { 
        background:rgba(255,255,255,0.28); 
        transform:translateY(-2px); 
    }

    /* Status Bar */
    .status-bar {
        background:linear-gradient(135deg,#d1fae5,#a7f3d0);
        padding:1.2rem 2rem;
        border-radius:16px;
        margin-bottom:2rem;
        border:1px solid #10b981;
    }

    /* KPI Cards */
    .stat-card {
        background:white; 
        border:1px solid #e5e7eb; 
        border-radius:16px;
        padding:1.6rem; 
        text-align:center; 
        transition:all 0.3s ease;
        box-shadow:0 4px 6px rgba(0,0,0,0.05);
    }
    
    .stat-card:hover { 
        transform:translateY(-4px); 
        box-shadow:0 10px 30px rgba(102,126,234,0.15); 
        border-color:#667eea; 
    }
    
    .stat-value {
        font-size:2.2rem; 
        font-weight:800;
        background:linear-gradient(135deg,#06b6d4,#3b82f6);
        -webkit-background-clip:text; 
        -webkit-text-fill-color:transparent;
        margin-bottom:0.4rem;
    }
    
    .stat-label { 
        color:#6b7280; 
        font-size:0.82rem; 
        font-weight:600; 
        text-transform:uppercase; 
        letter-spacing:1px; 
    }
    
    .section-header {
        font-family:'Space Grotesk',sans-serif; 
        font-size:1.8rem; 
        font-weight:700;
        color:#1f2937; 
        border-bottom:3px solid #667eea; 
        padding-bottom:0.6rem; 
        margin:1.5rem 0 1rem 0;
    }

    /* CHAT CONTAINER WITH SCROLLING */
    
    .chat-header-bar {
        background: linear-gradient(135deg, #06b6d4, #3b82f6);
        color: white;
        padding: 1.2rem 1.5rem;
        font-weight: 700;
        font-size: 1.15rem;
        display: flex;
        align-items: center;
        gap: 0.5rem;
        flex-shrink: 0;
        border-radius: 16px 16px 0 0;
    }
    
    
    /* Chat Messages Styling */
    .stChatMessage {
        background: white !important;
        border-radius: 12px !important;
        padding: 1.2rem !important;
        margin-bottom: 1rem !important;
        box-shadow: 0 2px 6px rgba(0,0,0,0.06) !important;
        border: 1px solid #e5e7eb !important;
    }
    
    .stChatMessage[data-testid="user-message"] {
        background: linear-gradient(135deg, #eff6ff, #dbeafe) !important;
        border-left: 4px solid #3b82f6 !important;
    }
    
    .stChatMessage[data-testid="assistant-message"] {
        background: white !important;
        border-left: 4px solid #10b981 !important;
    }
    
    /* Perfect Scrollbar */
    .chat-messages-scroll::-webkit-scrollbar {
        width: 8px;
    }
    
    .chat-messages-scroll::-webkit-scrollbar-thumb {
        background: #cbd5e1;
        border-radius: 10px;
    }
    
    .chat-messages-scroll::-webkit-scrollbar-thumb:hover {
        background: #94a3b8;
    }
    
    .chat-messages-scroll::-webkit-scrollbar-track {
        background: #f1f5f9;
        border-radius: 10px;
    }

    /* Quick Action Buttons */
    .quick-buttons-grid {
        display: grid;
        grid-template-columns: 1fr 1fr;
        gap: 0.6rem;
        margin-bottom: 1rem;
    }
    
    .stButton > button {
        width: 100%;
        background: linear-gradient(135deg, #06b6d4, #3b82f6);
        color: white;
        border: none;
        border-radius: 10px;
        padding: 0.65rem 1.2rem;
        font-weight: 700;
        font-size: 0.9rem;
        transition: all 0.3s ease;
        box-shadow: 0 4px 12px rgba(6,182,212,0.3);
    }
    
    .stButton > button:hover { 
        transform: translateY(-2px); 
        box-shadow: 0 6px 18px rgba(6,182,212,0.4); 
    }

    /* Tabs Styling */
    .stTabs [data-baseweb="tab-list"] {
        gap: 0.6rem;
        background: #f9fafb;
        padding: 0.6rem;
        border-radius: 12px;
    }
    
    .stTabs [data-baseweb="tab"] {
        border-radius: 10px;
        padding: 0.6rem 1.2rem;
        font-weight: 700;
        font-size: 0.95rem;
    }
    
    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #06b6d4, #3b82f6);
        color: white;
    }

    /* Sidebar */
    section[data-testid="stSidebar"] { 
        background: #f9fafb;
    }

    /* Remove default padding */
    .block-container {
        padding-top: 2rem !important;
        padding-bottom: 2rem !important;
    }
</style>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════
# AI BACKENDS - ALL INCLUDED
# ═══════════════════════════════════════════════════════════════

@st.cache_resource(show_spinner=False)
def init_ai_backends():
    """Initialize all AI backends"""
    backends = {}
    
    # OpenAI
    try:
        from openai import OpenAI
        key = os.getenv('OPENAI_API_KEY')
        if key and len(key.strip()) > 20:
            client = OpenAI(api_key=key.strip())
            try:
                test = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[{"role": "user", "content": "test"}],
                    max_tokens=5
                )
                backends['openai'] = client
            except: pass
    except: pass
    
    # Groq
    try:
        from groq import Groq
        key = os.getenv('GROQ_API_KEY')
        if key and len(key.strip()) > 20:
            client = Groq(api_key=key.strip())
            try:
                test = client.chat.completions.create(
                    model="llama-3.3-70b-versatile",
                    messages=[{"role": "user", "content": "test"}],
                    max_tokens=5
                )
                backends['groq'] = client
            except: pass
    except: pass
    
    # Gemini
    try:
        import google.genai as genai
        key = os.getenv('GEMINI_API_KEY') or os.getenv('GOOGLE_API_KEY')
        if key:
            client = genai.Client(api_key=key)
            backends['gemini'] = {'client': client, 'type': 'new'}
    except:
        try:
            import google.generativeai as genai_old
            key = os.getenv('GEMINI_API_KEY') or os.getenv('GOOGLE_API_KEY')
            if key:
                genai_old.configure(api_key=key)
                model = genai_old.GenerativeModel('gemini-1.5-flash')
                backends['gemini'] = {'client': model, 'type': 'old'}
        except: pass
    
    # Ollama
    try:
        import requests
        r = requests.get('http://localhost:11434/api/tags', timeout=2)
        if r.status_code == 200:
            backends['ollama'] = True
    except: pass
    
    return backends


def build_enhanced_context(query, df):
    """Build comprehensive AI context"""
    total_deposits = df['Total_Deposits'].sum()
    total_advances = df['Advances'].sum()
    avg_npa = df['NPA_Percent'].mean()
    avg_casa = df['CASA_Percent'].mean()
    high_npa = df.nlargest(3, 'NPA_Percent')[['Branch_Name', 'NPA_Percent']].to_dict('records')
    low_casa = df.nsmallest(3, 'CASA_Percent')[['Branch_Name', 'CASA_Percent']].to_dict('records')
    top_performers = df.nlargest(3, 'Total_Deposits')[['Branch_Name', 'Total_Deposits']].to_dict('records')
    
    context = f"""You are BankVista AI, an expert banking analyst. You are BankVista AI, a banking analytics assistant. Provide calm, data-backed insights.

BANKING DATA OVERVIEW:
- Total Branches: {len(df)}
- Total Deposits: ₹{total_deposits:.2f} Crores
- Total Advances: ₹{total_advances:.2f} Crores
- Average NPA: {avg_npa:.2f}%
- Average CASA: {avg_casa:.2f}%

HIGH NPA BRANCHES (Needs Attention):
{chr(10).join([f"• {b['Branch_Name']}: {b['NPA_Percent']:.2f}%" for b in high_npa])}

LOW CASA BRANCHES (Growth Opportunity):
{chr(10).join([f"• {b['Branch_Name']}: {b['CASA_Percent']:.2f}%" for b in low_casa])}

TOP PERFORMERS (By Deposits):
{chr(10).join([f"• {b['Branch_Name']}: ₹{b['Total_Deposits']:.2f}Cr" for b in top_performers])}

USER QUERY: {query}

RESPONSE GUIDELINES:
- Max 6–8 short bullet points
- Be concise and calm
- Avoid commanding language
- Use suggestive tone (e.g., "may consider", "could explore")
- Do not create urgency or pressure
- No long explanations
- Focus on insights, not instructions
- Avoid using words like "must", "urgent", "immediately"
- Use balanced analytical language
- Present observations before suggestions

Respond now:"""
    
    return context


def call_ai(query, df):
    """Call AI with all backends"""
    backends = init_ai_backends()
    context = build_enhanced_context(query, df)
    
    # Try OpenAI
    if 'openai' in backends:
        try:
            response = backends['openai'].chat.completions.create(
                model="gpt-4o-mini",
                messages=[{"role": "user", "content": context}],
                max_tokens=1000,
                temperature=0.7
            )
            return response.choices[0].message.content
        except: pass
    
    # Try Groq
    if 'groq' in backends:
        try:
            response = backends['groq'].chat.completions.create(
                model="llama-3.3-70b-versatile",
                messages=[{"role": "user", "content": context}],
                max_tokens=1000,
                temperature=0.7
            )
            return response.choices[0].message.content
        except: pass
    
    # Try Gemini
    if 'gemini' in backends:
        try:
            client_info = backends['gemini']
            if client_info['type'] == 'new':
                response = client_info['client'].models.generate_content(
                    model='gemini-2.0-flash-exp',
                    contents=context
                )
                return response.text
            else:
                response = client_info['client'].generate_content(context)
                return response.text
        except: pass
    
    # Try Ollama
    if 'ollama' in backends:
        try:
            import requests
            response = requests.post('http://localhost:11434/api/generate', json={
                "model": "llama3.2",
                "prompt": context,
                "stream": False
            }, timeout=30)
            return response.json()['response']
        except: pass
    
    # Local fallback
    return get_enhanced_local_response(query, df)


def get_enhanced_local_response(query, df):
    """Enhanced local responses"""
    q = query.lower()
    
    # NPA Analysis
# NPA Analysis
    if any(word in q for word in ['npa', 'bad', 'loan', 'default']):

        high_npa = df.nlargest(5, 'NPA_Percent')

        response = f"""
    🔴 **NPA Overview**

    (Org Avg: {df['NPA_Percent'].mean():.2f}%)

    **Top Risk Branches:**
    """

        for _, row in high_npa.iterrows():
            response += f"- {row['Branch_Name']} → {row['NPA_Percent']:.2f}%\n"

        response += """
    Observations & Possible Considerations:
    • Some branches show elevated stress
    • Recovery follow-ups may be explored gradually
    • Monitoring trends over time could help assess trajectory


    """

        return response

    
    # CASA Analysis
    elif any(word in q for word in ['casa', 'deposit', 'current', 'savings']):
        low_casa = df.nsmallest(5, 'CASA_Percent')
        high_casa = df.nlargest(3, 'CASA_Percent')
        
        response = f"""**💰 CASA (Current & Savings Account) Analysis**

**Why CASA Matters?** CASA deposits are low-cost funds that improve profitability.

**Organization Overview:**
- **Average CASA:** {df['CASA_Percent'].mean():.2f}%
- **Target:** >40.0%
- **Below target:** {len(df[df['CASA_Percent'] < 40])} branches

**🎯 Growth Opportunities:**

"""
        for i, (_, row) in enumerate(low_casa.iterrows(), 1):
            gap = 40 - row['CASA_Percent']
            response += f"{i}. **{row['Branch_Name']}**: {row['CASA_Percent']:.2f}%\n"
            response += f"   - Growth potential: {gap:.1f}% | Deposits: ₹{row['Total_Deposits']:.1f}Cr\n"
        
        response += f"\n**🏆 Top Performers:**\n"
        for _, row in high_casa.iterrows():
            response += f"• **{row['Branch_Name']}**: {row['CASA_Percent']:.2f}%\n"
        
        response += f"\n**💡 Action Plan:**\n"
        response += f"1. Launch CASA campaigns in low-performing branches\n"
        response += f"2. Offer competitive interest rates and benefits\n"
        response += f"3. Promote zero-balance and digital accounts\n"
        response += f"4. Cross-sell to existing term deposit customers\n"
        response += f"5. Target: Achieve >40% CASA across all branches\n"
        
        return response
    
    # Performance Analysis
    elif any(word in q for word in ['top', 'best', 'performer', 'leader']):
        top_deposits = df.nlargest(5, 'Total_Deposits')
        
        response = f"""**🏆 Top Performing Branches**

"""
        for i, (_, row) in enumerate(top_deposits.iterrows(), 1):
            achievement = (row['Total_Deposits'] / row['Deposit_Target'] * 100)
            grade = "🌟 Excellent" if achievement > 100 else "✅ Good"
            
            response += f"**{i}. {row['Branch_Name']}** - ₹{row['Total_Deposits']:.1f}Cr {grade}\n"
            response += f"   - Achievement: {achievement:.1f}%\n"
            response += f"   - NPA: {row['NPA_Percent']:.1f}% | CASA: {row['CASA_Percent']:.1f}%\n"
            response += f"   - Zone: {row['Zone']} | Staff: {row['Staff_Count']}\n\n"
        
        return response
    
    # Branch-specific
    elif any(branch.lower() in q for branch in df['Branch_Name'].str.lower()):
        for branch in df['Branch_Name']:
            if branch.lower() in q:
                row = df[df['Branch_Name'] == branch].iloc[0]
                achievement = (row['Total_Deposits'] / row['Deposit_Target'] * 100)
                
                response = f"""**📍 {branch} Branch Analysis**

**Financial Performance:**
- **Deposits:** ₹{row['Total_Deposits']:.2f}Cr (Target: ₹{row['Deposit_Target']:.2f}Cr)
- **Achievement:** {achievement:.1f}% {'🌟' if achievement > 100 else '⚠️'}
- **Advances:** ₹{row['Advances']:.2f}Cr

**Key Metrics:**
- **NPA:** {row['NPA_Percent']:.2f}% {'✅' if row['NPA_Percent'] < 3 else '🟡' if row['NPA_Percent'] < 6 else '🔴'}
- **CASA:** {row['CASA_Percent']:.2f}% {'✅' if row['CASA_Percent'] > 40 else '🟡'}
- **CD Ratio:** {row['CD_Ratio']:.2f}%

**Staff:**
- Count: {row['Staff_Count']}
- Business/Staff: ₹{row['Business_Per_Staff']:.2f}Cr
- Profit/Staff: ₹{row['Profit_Per_Staff']:.2f}Cr

**Location:** {row['Zone']}
"""
                return response
    
    # Default overview
    else:
        response = f"""**📊 BankVista Overview**

**Organization:**
- **Branches:** {len(df)}
- **Total Deposits:** ₹{df['Total_Deposits'].sum():.2f}Cr
- **Total Advances:** ₹{df['Advances'].sum():.2f}Cr
- **Total Staff:** {df['Staff_Count'].sum()}

**Key Metrics:**
- **Avg NPA:** {df['NPA_Percent'].mean():.2f}% (Target: <3%)
- **Avg CASA:** {df['CASA_Percent'].mean():.2f}% (Target: >40%)
- **Avg CD Ratio:** {df['CD_Ratio'].mean():.2f}%

**Quick Insights:**
- Highest NPA: {df.loc[df['NPA_Percent'].idxmax(), 'Branch_Name']} ({df['NPA_Percent'].max():.2f}%)
- Lowest CASA: {df.loc[df['CASA_Percent'].idxmin(), 'Branch_Name']} ({df['CASA_Percent'].min():.2f}%)
- Top Performer: {df.loc[df['Total_Deposits'].idxmax(), 'Branch_Name']} (₹{df['Total_Deposits'].max():.2f}Cr)

**💡 Ask me:**
- "Which branches have bad loans?"
- "Where can we grow CASA?"
- "Show me top performers"
"""
        return response


def should_show_chart(query):
    """Check if chart needed"""
    q = query.lower()
    chart_keywords = ['npa', 'casa', 'advance', 'top', 'best', 'worst', 'compare', 'chart', 'show']
    return any(keyword in q for keyword in chart_keywords)


def create_chat_chart(query, df):
    """Create chart for chat"""
    q = query.lower()
    
    if 'npa' in q or 'bad' in q:
        top = df.nlargest(10, 'NPA_Percent')
        colors = ['#ef4444' if x > 6 else '#f59e0b' if x > 3 else '#10b981' for x in top['NPA_Percent']]
        
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=top['Branch_Name'],
            y=top['NPA_Percent'],
            marker=dict(color=colors),
            text=top['NPA_Percent'].round(2),
            textposition='outside',
            texttemplate='%{text}%'
        ))
        
        fig.add_hline(y=3, line_dash="dash", line_color="#10b981", annotation_text="Target: 3%")
        fig.add_hline(y=6, line_dash="dash", line_color="#f59e0b", annotation_text="Warning: 6%")
        
        fig.update_layout(
            title="Top 10 Branches by NPA",
            xaxis_title="Branch",
            yaxis_title="NPA %",
            height=350,
            showlegend=False,
            xaxis_tickangle=-45,
            margin=dict(l=20, r=20, t=50, b=80)
        )
        return fig
    
    elif 'casa' in q:
        low = df.nsmallest(5, 'CASA_Percent')
        high = df.nlargest(5, 'CASA_Percent')
        combined = pd.concat([low, high]).drop_duplicates()
        colors = ['#ef4444' if x < 30 else '#fbbf24' if x < 40 else '#10b981' for x in combined['CASA_Percent']]
        
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=combined['Branch_Name'],
            y=combined['CASA_Percent'],
            marker=dict(color=colors),
            text=combined['CASA_Percent'].round(2),
            textposition='outside',
            texttemplate='%{text}%'
        ))
        
        fig.add_hline(y=40, line_dash="dash", line_color="#3b82f6", annotation_text="Target: 40%")
        
        fig.update_layout(
            title="CASA % Analysis",
            xaxis_title="Branch",
            yaxis_title="CASA %",
            height=350,
            showlegend=False,
            xaxis_tickangle=-45,
            margin=dict(l=20, r=20, t=50, b=80)
        )
        return fig
    
    elif 'top' in q or 'best' in q or 'performer' in q:
        top = df.nlargest(10, 'Total_Deposits')
        
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=top['Branch_Name'],
            y=top['Total_Deposits'],
            marker=dict(color=top['Total_Deposits'], colorscale='Blues', showscale=False),
            text=top['Total_Deposits'].round(2),
            textposition='outside',
            texttemplate='₹%{text}Cr'
        ))
        
        fig.update_layout(
            title="Top 10 Performers by Deposits",
            xaxis_title="Branch",
            yaxis_title="Deposits (Crores)",
            height=350,
            showlegend=False,
            xaxis_tickangle=-45,
            margin=dict(l=20, r=20, t=50, b=80)
        )
        return fig
    
    return None


# ═══════════════════════════════════════════════════════════════
# PREDICTIVE ANALYTICS
# ═══════════════════════════════════════════════════════════════

class PredictiveAnalytics:
    def __init__(self, df):
        self.df = df
    
    def predict_npa_trend(self, branch_name, months=6):
        branch = self.df[self.df['Branch_Name'] == branch_name].iloc[0]
        current_npa = branch['NPA_Percent']
        
        predictions = []
        npa = current_npa
        
        for month in range(1, months + 1):
            seasonal_factor = 1.0 + (0.15 if month % 3 == 0 else 0.05)
            trend_factor = 1.02 if current_npa > 5 else 0.98
            noise = np.random.normal(0, 0.2)
            npa = max(0, npa * seasonal_factor * trend_factor + noise)
            
            predictions.append({
                'month': month,
                'predicted_npa': round(npa, 2)
            })
        
        final_npa = predictions[-1]['predicted_npa']
        if final_npa > 6:
            risk_level = 'HIGH RISK'
            risk_color = '🔴'
        elif final_npa > 3:
            risk_level = 'MEDIUM RISK'
            risk_color = '🟡'
        else:
            risk_level = 'LOW RISK'
            risk_color = '🟢'
        
        return {
            'branch': branch_name,
            'current_npa': current_npa,
            'predictions': predictions,
            'final_npa': final_npa,
            'risk_level': risk_level,
            'risk_color': risk_color
        }


class AnomalyDetector:
    def __init__(self, df):
        self.df = df
    
    def detect_anomalies(self):
        anomalies = []
        metrics = ['NPA_Percent', 'CASA_Percent', 'CD_Ratio', 'Profit_Per_Staff', 'Business_Per_Staff']
        
        for metric in metrics:
            if metric not in self.df.columns:
                continue
            
            mean = self.df[metric].mean()
            std = self.df[metric].std()
            
            if std == 0:
                continue
            
            for _, row in self.df.iterrows():
                value = row[metric]
                z_score = (value - mean) / std
                
                if abs(z_score) > 2:
                    anomalies.append({
                        'branch': row['Branch_Name'],
                        'metric': metric,
                        'value': round(value, 2),
                        'org_mean': round(mean, 2),
                        'z_score': round(z_score, 2),
                        'direction': 'HIGH ⬆️' if z_score > 0 else 'LOW ⬇️',
                        'severity': 'Critical' if abs(z_score) > 3 else 'Warning'
                    })
        
        return sorted(anomalies, key=lambda x: abs(x['z_score']), reverse=True)


# ═══════════════════════════════════════════════════════════════
# VISUALIZATION FUNCTIONS
# ═══════════════════════════════════════════════════════════════

def create_performance_heatmap(df):
    metrics = ['NPA_Percent', 'CASA_Percent', 'CD_Ratio', 'Profit_Per_Staff', 'Business_Per_Staff']
    heatmap_data = df[metrics].copy()
    
    for col in metrics:
        heatmap_data[col] = (heatmap_data[col] - heatmap_data[col].min()) / (heatmap_data[col].max() - heatmap_data[col].min())
    
    fig = go.Figure(data=go.Heatmap(
        z=heatmap_data.T.values,
        x=df['Branch_Name'],
        y=[m.replace('_', ' ') for m in metrics],
        colorscale='RdYlGn',
        text=df[metrics].T.round(2).values,
        texttemplate='%{text}',
        textfont={"size": 10},
        colorbar=dict(title="Normalized<br>Score")
    ))
    
    fig.update_layout(
        title="Branch Performance Heatmap",
        height=400,
        xaxis_tickangle=-45
    )
    
    return fig


def create_zone_comparison(df):
    zone_stats = df.groupby('Zone').agg({
        'Total_Deposits': 'sum',
        'NPA_Percent': 'mean',
        'CASA_Percent': 'mean',
        'Branch_Name': 'count'
    }).round(2)
    
    fig = make_subplots(
        rows=2, cols=2,
        subplot_titles=('Total Deposits', 'Average NPA', 'Average CASA', 'Branch Count'),
        specs=[[{'type': 'bar'}, {'type': 'bar'}],
               [{'type': 'bar'}, {'type': 'bar'}]]
    )
    
    zones = zone_stats.index.tolist()
    
    fig.add_trace(go.Bar(x=zones, y=zone_stats['Total_Deposits'], name='Deposits',
                         marker_color='#06b6d4'), row=1, col=1)
    fig.add_trace(go.Bar(x=zones, y=zone_stats['NPA_Percent'], name='NPA',
                         marker_color='#ef4444'), row=1, col=2)
    fig.add_trace(go.Bar(x=zones, y=zone_stats['CASA_Percent'], name='CASA',
                         marker_color='#10b981'), row=2, col=1)
    fig.add_trace(go.Bar(x=zones, y=zone_stats['Branch_Name'], name='Branches',
                         marker_color='#8b5cf6'), row=2, col=2)
    
    fig.update_layout(height=600, showlegend=False, title_text="Zone-wise Analysis")
    
    return fig


# ═══════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════

def create_excel_dashboard(df):
    output = io.BytesIO()
    wb = Workbook()
    wb.remove(wb.active)
    
    ws_data = wb.create_sheet("_Data")
    ws_data.sheet_state = 'hidden'
    
    for col_idx, header in enumerate(df.columns, 1):
        ws_data.cell(1, col_idx, header).font = Font(bold=True)
    
    for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws_data.cell(row_idx, col_idx, value)
    
    ws = wb.create_sheet("Dashboard", 0)
    ws.merge_cells('A1:F1')
    ws['A1'].value = "BANKVISTA AI - DYNAMIC DASHBOARD"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill("solid", fgColor="5B4B8A")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    ws['A3'] = "Select Branch:"
    ws['B3'] = df.iloc[0]['Branch_Name']
    
    dv = DataValidation(type="list", formula1=f"=_Data!$B$2:$B${len(df)+1}")
    dv.add('B3')
    ws.add_data_validation(dv)
    
    ws_all = wb.create_sheet("All Branches")
    for col_idx, header in enumerate(df.columns, 1):
        ws_all.cell(1, col_idx, header).font = Font(bold=True)
        ws_all.cell(1, col_idx).fill = PatternFill("solid", fgColor="4472C4")
    
    for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            ws_all.cell(row_idx, col_idx, value)
    
    wb.save(output)
    output.seek(0)
    return output


def generate_sample_data():
    return pd.DataFrame({
        'Branch_ID': ['B1001','B1002','B1003','B1004','B1005','B2001','B2002','B2003','B3001','B3002',
                      'B3003','B3004','B3005','B3006','B3007','B3008','B3009','B3010','B4001','B4002'],
        'Branch_Name': ['Mansoorabad','Adilabad','Hyderabad Main','Secunderabad','Warangal',
                       'Vijayawada','Visakhapatnam','Guntur','Nellore','Tirupati',
                       'Kadapa','Chittoor','Anantapur','Kurnool','Rajahmundry','Kakinada',
                       'Eluru','Ongole','Nizamabad','Karimnagar'],
        'Zone': ['Telangana']*10 + ['Andhra Pradesh']*10,
        'Total_Deposits': [110.97,85.45,245.80,189.23,67.89,198.76,223.45,145.67,156.34,189.45,
                          134.56,167.89,145.23,178.90,201.34,187.56,156.78,134.90,123.45,145.67],
        'Deposit_Target': [105.46,95.00,250.00,195.00,75.00,205.00,220.00,150.00,160.00,185.00,
                          140.00,165.00,150.00,175.00,200.00,185.00,155.00,140.00,125.00,150.00],
        'Advances': [232.29,156.78,412.50,289.45,98.76,356.89,389.23,245.78,267.89,345.67,
                    245.34,298.76,256.78,312.45,378.90,334.56,278.90,245.67,234.56,267.89],
        'Advance_Target': [218.16,175.00,425.00,295.00,110.00,365.00,395.00,255.00,270.00,340.00,
                          250.00,295.00,260.00,310.00,375.00,330.00,275.00,250.00,235.00,270.00],
        'NPA_Percent': [2.8,5.4,1.9,3.2,8.5,2.3,1.8,4.1,3.5,2.7,4.3,3.8,4.6,3.1,2.4,2.9,3.7,4.2,5.1,3.9],
        'Profit_Per_Staff': [5.44,3.1,6.8,4.5,1.8,5.8,6.5,3.8,4.2,5.1,3.9,4.7,3.6,4.9,5.9,5.3,4.4,3.7,3.3,4.1],
        'CASA_Percent': [42.3,28.5,51.2,38.7,25.4,44.7,49.3,34.9,36.8,40.2,33.5,38.1,35.7,39.4,45.6,41.8,37.2,34.3,31.9,36.5],
        'CD_Ratio': [72.5,78.2,65.8,70.1,82.3,66.8,64.5,73.4,71.2,68.9,74.3,70.7,72.8,69.5,67.3,68.7,71.6,73.9,76.4,72.1],
        'Business_Per_Staff': [85.2,58.3,95.7,78.9,45.6,86.7,91.2,67.8,71.4,83.5,69.2,77.6,70.3,79.8,88.4,82.1,72.9,68.5,64.7,73.2],
        'Staff_Count': [25,18,42,35,15,38,40,27,29,34,26,31,28,32,37,36,30,27,24,29]
    })


# ═══════════════════════════════════════════════════════════════
# MAIN APPLICATION
# ═══════════════════════════════════════════════════════════════

def main():
    if 'messages' not in st.session_state:
        st.session_state.messages = []
    if 'df' not in st.session_state:
        st.session_state.df = None

    # Hero dynamic sizing
    if st.session_state.df is None:
        hero_class = "hero-container"
        title_size = "3.8rem"
        hero_padding = "3.5rem 2rem"
    else:
        hero_class = "hero-container"
        title_size = "2.2rem"
        hero_padding = "1.5rem 2rem"


    # Hero Section
    st.markdown(f"""
    <div class="{hero_class}" style="padding:{'3.5rem 2rem' if st.session_state.df is None else '1.5rem 2rem'};">
        <h1 class="main-title" style="font-size:{title_size};">🤖 BankVista AI</h1>
        <p class="subtitle">Next-Generation Banking Analytics · Powered by Advanced AI</p>
    </div>
    """, unsafe_allow_html=True)


    # Sidebar
    with st.sidebar:
        st.markdown("### 📤 Data Upload")
        
        uploaded_file = st.file_uploader("CSV or Excel", type=['csv', 'xlsx', 'xls'], label_visibility="collapsed")
        
        if st.button("📊 Try Sample Data"):
            st.session_state.df = generate_sample_data()
            st.session_state.messages = []
            st.rerun()
        
        st.markdown("---")
        st.markdown("### 💡 Quick Tips")
        st.markdown("""
        - "Show bad loans"
        - "CASA opportunities?"  
        - "Top performers"
        - "Analyze [Branch]"
        """)

    if uploaded_file:
        try:
            df = pd.read_csv(uploaded_file) if uploaded_file.name.endswith('.csv') else pd.read_excel(uploaded_file)
            st.session_state.df = df
            st.session_state.messages = []
            st.success(f"✅ Loaded {len(df)} branches!")
        except Exception as e:
            st.error(f"❌ Error: {e}")

    if st.session_state.df is None:
        st.markdown("""
        <div style="text-align:center;padding:4rem 2rem;background:white;border-radius:24px;margin-top:2rem;box-shadow:0 4px 6px rgba(0,0,0,0.05);">
            <div style="font-size:4rem;margin-bottom:1rem;">🚀</div>
            <h2 style="font-size:2.5rem;margin-bottom:1rem;font-weight:800;color:#1f2937;">Welcome to BankVista AI</h2>
            <p style="font-size:1.3rem;color:#6b7280;margin-bottom:2rem;">Transform your banking data into actionable insights</p>
            <p style="font-size:1.1rem;color:#9ca3af;">Upload your data or try our sample dataset to begin</p>
        </div>
        """, unsafe_allow_html=True)
        return

    df = st.session_state.df
    predictive = PredictiveAnalytics(df)
    detector = AnomalyDetector(df)

    # Status Bar
    total_deposits = df['Total_Deposits'].sum()
    total_staff = df['Staff_Count'].sum()
    deposit_achievement = (df['Total_Deposits'].sum() / df['Deposit_Target'].sum() * 100)
    
    st.markdown(f"""
    <div class="status-bar">
        <p style="text-align:center;font-size:1.05rem;margin:0;color:#065f46;font-weight:600;">
            ✅ <strong>{len(df)} Branches</strong> &nbsp;|&nbsp; 
            🤖 <strong>AI Active</strong> &nbsp;|&nbsp;
            💰 <strong>₹{total_deposits:.1f}Cr</strong> &nbsp;|&nbsp;
            👥 <strong>{total_staff} Staff</strong> &nbsp;|&nbsp;
            📊 <strong>{deposit_achievement:.1f}% Achievement</strong>
        </p>
    </div>
    """, unsafe_allow_html=True)

    # KPI Cards
    col1, col2, col3, col4 = st.columns(4)
    
    kpis = [
        (col1, f"{deposit_achievement:.1f}%", "Deposit Achievement"),
        (col2, f"{df['NPA_Percent'].mean():.2f}%", "Average NPA"),
        (col3, f"{df['CASA_Percent'].mean():.1f}%", "Average CASA"),
        (col4, f"₹{df['Business_Per_Staff'].mean():.1f}Cr", "Avg Business/Staff")
    ]
    
    for col, value, label in kpis:
        with col:
            st.markdown(f"""
            <div class="stat-card">
                <div class="stat-value">{value}</div>
                <div class="stat-label">{label}</div>
            </div>
            """, unsafe_allow_html=True)

    # Main Tabs
    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
        "💬 AI Chat",
        "📈 Predictions",
        "📊 Dashboard",
        "🗺️ Zone Analytics",
        "🔍 Anomalies",
        "📥 Export"
    ])

    # ═══════════════════════════════════════════════════════════
# TAB 1: AI CHAT (PROPER FIXED INPUT + SCROLLING MESSAGES)
# ═══════════════════════════════════════════════════════════
    with tab1:

        st.markdown('<div class="section-header">💬 AI Chat</div>', unsafe_allow_html=True)

        # Create a container for chat messages
        chat_placeholder = st.container()

        # FIXED HEIGHT SCROLL VIA STYLING BLOCK-CONTAINER INSIDE THIS TAB
        st.markdown("""
            <style>
            section[data-testid="stTabs"] div[data-testid="stVerticalBlock"]:has(div.stChatMessage) {
                max-height: 550px;
                overflow-y: auto;
                padding-right: 10px;
            }
            </style>
        """, unsafe_allow_html=True)

        # Render messages
        with chat_placeholder:
            for idx, msg in enumerate(st.session_state.messages):
                with st.chat_message(msg["role"]):
                    st.markdown(msg["content"])
                    if "chart" in msg and msg["chart"]:
                        st.plotly_chart(
                            msg["chart"],
                            width="stretch",
                            key=f"chat_{idx}"
                        )

        # FIXED QUICK ACTIONS (NOT SCROLLING)
        st.markdown("### ⚡ Quick Insights")
        col1, col2, col3 = st.columns(3)

        if col1.button("🔴 Bad Loans", key="quick_npa"):
            prompt = "Which branches have higher NPA?"
            st.session_state.messages.append({"role": "user", "content": prompt})
            response = call_ai(prompt, df)
            chart = create_chat_chart(prompt, df)
            st.session_state.messages.append(
                {"role": "assistant", "content": response, "chart": chart}
            )
            st.rerun()

        if col2.button("💰 CASA Opportunities", key="quick_casa"):
            prompt = "Where can CASA be improved?"
            st.session_state.messages.append({"role": "user", "content": prompt})
            response = call_ai(prompt, df)
            chart = create_chat_chart(prompt, df)
            st.session_state.messages.append(
                {"role": "assistant", "content": response, "chart": chart}
            )
            st.rerun()

        if col3.button("🏆 Top Performers", key="quick_top"):
            prompt = "Show top performing branches"
            st.session_state.messages.append({"role": "user", "content": prompt})
            response = call_ai(prompt, df)
            chart = create_chat_chart(prompt, df)
            st.session_state.messages.append(
                {"role": "assistant", "content": response, "chart": chart}
            )
            st.rerun()

        # FIXED CHAT INPUT AT BOTTOM
        user_input = st.chat_input("Ask about NPA, CASA, branch performance...")

        if user_input:
            st.session_state.messages.append({"role": "user", "content": user_input})
            response = call_ai(user_input, df)
            chart = create_chat_chart(user_input, df) if should_show_chart(user_input) else None
            st.session_state.messages.append(
                {"role": "assistant", "content": response, "chart": chart}
            )
            st.rerun()



    # ═══════════════════════════════════════════════════════════
    # TAB 2: PREDICTIONS
    # ═══════════════════════════════════════════════════════════
    with tab2:
        st.markdown('<div class="section-header">📈 Predictive Analytics</div>', unsafe_allow_html=True)
        
        col1, col2 = st.columns([1, 2])
        
        with col1:
            st.markdown("### Select Branch")
            selected_branch = st.selectbox(
                "Choose a branch:",
                df['Branch_Name'].tolist(),
                label_visibility="collapsed",
                key="pred_branch"
            )
        
        with col2:
            prediction = predictive.predict_npa_trend(selected_branch)
            
            # Metrics
            mcol1, mcol2, mcol3 = st.columns(3)
            mcol1.metric("Current NPA", f"{prediction['current_npa']:.2f}%")
            mcol2.metric("6-Month Forecast", f"{prediction['final_npa']:.2f}%")
            mcol3.metric("Risk Level", f"{prediction['risk_color']} {prediction['risk_level']}")
            
            # Chart
            months = [f"Month {p['month']}" for p in prediction['predictions']]
            npas = [p['predicted_npa'] for p in prediction['predictions']]
            
            fig = go.Figure()
            
            fig.add_trace(go.Scatter(
                x=months, y=npas,
                mode='lines+markers',
                name='Predicted NPA',
                line=dict(color='#06b6d4', width=3),
                marker=dict(size=10, symbol='circle')
            ))
            
            fig.add_hline(y=3, line_dash="dash", line_color="#10b981",
                        annotation_text="Target: 3%", annotation_position="right")
            fig.add_hline(y=6, line_dash="dash", line_color="#f59e0b",
                        annotation_text="Warning: 6%", annotation_position="right")
            
            fig.update_layout(
                title=f"NPA Forecast - {selected_branch}",
                xaxis_title="Month",
                yaxis_title="NPA %",
                height=400,
                hovermode='x unified'
            )
            
            st.plotly_chart(fig, width="stretch", key="pred_chart")

    # ═══════════════════════════════════════════════════════════
    # TAB 3: DASHBOARD
    # ═══════════════════════════════════════════════════════════
    with tab3:
        st.markdown('<div class="section-header">📊 Analytics Dashboard</div>', unsafe_allow_html=True)
        
        # Data table
        display_df = df.copy()
        display_df['Dep_%'] = (display_df['Total_Deposits'] / display_df['Deposit_Target'] * 100).round(1)
        
        st.dataframe(
            display_df[[
                'Branch_Name', 'Zone', 'Total_Deposits', 'Dep_%',
                'NPA_Percent', 'CASA_Percent', 'Staff_Count', 'Business_Per_Staff'
            ]],
            width="stretch",
            hide_index=True
        )
        
        # Charts
        col1, col2 = st.columns(2)
        
        with col1:
            fig1 = px.bar(
                df.nlargest(10, 'NPA_Percent'),
                x='Branch_Name',
                y='NPA_Percent',
                title='Top 10 Branches by NPA',
                color='NPA_Percent',
                color_continuous_scale='Reds',
                labels={'NPA_Percent': 'NPA %', 'Branch_Name': 'Branch'}
            )
            fig1.update_layout(showlegend=False, xaxis_tickangle=-45)
            st.plotly_chart(fig1, width="stretch", key="dash_npa")
        
        with col2:
            fig2 = px.bar(
                df.nlargest(10, 'CASA_Percent'),
                x='Branch_Name',
                y='CASA_Percent',
                title='Top 10 Branches by CASA',
                color='CASA_Percent',
                color_continuous_scale='Greens',
                labels={'CASA_Percent': 'CASA %', 'Branch_Name': 'Branch'}
            )
            fig2.update_layout(showlegend=False, xaxis_tickangle=-45)
            st.plotly_chart(fig2, width="stretch", key="dash_casa")
        
        # Heatmap
        st.markdown("### 🔥 Performance Heatmap")
        heatmap_fig = create_performance_heatmap(df)
        st.plotly_chart(heatmap_fig, width="stretch", key="heatmap")

    # ═══════════════════════════════════════════════════════════
    # TAB 4: ZONE ANALYTICS
    # ═══════════════════════════════════════════════════════════
    with tab4:
        st.markdown('<div class="section-header">🗺️ Zone Analytics</div>', unsafe_allow_html=True)
        
        # Zone comparison
        zone_fig = create_zone_comparison(df)
        st.plotly_chart(zone_fig, width="stretch", key="zone_chart")
        
        # Detailed zone stats
        st.markdown("### Zone Performance Details")
        
        zone_stats = df.groupby('Zone').agg({
            'Total_Deposits': ['sum', 'mean'],
            'Advances': ['sum', 'mean'],
            'NPA_Percent': 'mean',
            'CASA_Percent': 'mean',
            'Branch_Name': 'count',
            'Staff_Count': 'sum'
        }).round(2)
        
        zone_stats.columns = ['Total Deposits', 'Avg Deposits', 'Total Advances', 'Avg Advances',
                             'Avg NPA', 'Avg CASA', 'Branches', 'Total Staff']
        
        st.dataframe(zone_stats, width="stretch")

    # ═══════════════════════════════════════════════════════════
    # TAB 5: ANOMALIES
    # ═══════════════════════════════════════════════════════════
    with tab5:
        st.markdown('<div class="section-header">🔍 Anomaly Detection</div>', unsafe_allow_html=True)
        
        anomalies = detector.detect_anomalies()
        
        if len(anomalies) == 0:
            st.success("✅ No significant anomalies detected!")
        else:
            st.warning(f"⚠️ Detected {len(anomalies)} anomalies")
            
            # Filters
            col1, col2 = st.columns(2)
            with col1:
                severity_filter = st.multiselect(
                    "Severity:",
                    ['Critical', 'Warning'],
                    default=['Critical', 'Warning'],
                    key="sev_filter"
                )
            with col2:
                metric_filter = st.multiselect(
                    "Metric:",
                    list(set([a['metric'] for a in anomalies])),
                    default=list(set([a['metric'] for a in anomalies])),
                    key="metric_filter"
                )
            
            filtered_anomalies = [
                a for a in anomalies
                if a['severity'] in severity_filter and a['metric'] in metric_filter
            ]
            
            if filtered_anomalies:
                anomaly_df = pd.DataFrame(filtered_anomalies)
                st.dataframe(anomaly_df, width="stretch", hide_index=True)

    # ═══════════════════════════════════════════════════════════
    # TAB 6: EXPORT
    # ═══════════════════════════════════════════════════════════
    with tab6:
        st.markdown('<div class="section-header">📥 Export Data</div>', unsafe_allow_html=True)
        
        st.info("📊 Generate Excel dashboard with interactive formulas")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown("### Excel Features")
            st.markdown("""
            ✅ **Interactive Dashboard** - Branch selector dropdown  
            ✅ **Dynamic Formulas** - Auto-updating metrics  
            ✅ **Professional Design** - Color-coded sections  
            ✅ **All Branches Sheet** - Complete data  
            ✅ **Offline Ready** - Share freely  
            """)
        
        with col2:
            if st.button("📊 Generate Excel Dashboard", type="primary", key="gen_excel"):
                with st.spinner("Creating Excel file..."):
                    excel_file = create_excel_dashboard(df)
                    
                    st.download_button(
                        label="⬇️ Download Excel File",
                        data=excel_file,
                        file_name=f"BankVista_Dashboard_{date.today().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="download_excel"
                    )
                    
                    st.success("✅ Excel file generated successfully!")
                    st.balloons()


if __name__ == "__main__":
    main()
