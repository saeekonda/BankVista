from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import pandas as pd
from datetime import date
import os
import sys

green = PatternFill("solid", fgColor="C6EFCE")
amber = PatternFill("solid", fgColor="FFEB9C")
red = PatternFill("solid", fgColor="FFC7CE")

def colour_achievement(cell, value):
    if value >= 100:
        cell.fill = green
    elif value >= 90:
        cell.fill = amber
    else:
        cell.fill = red

def colour_gap(cell, value):
    if value <= 0:          # surplus
        cell.fill = green
    elif value <= 100:
        cell.fill = amber
    else:
        cell.fill = red

def generate_key_takeaways(b):
    insights = []

    # Deposits
    if b.total_deposits_cr >= b.deposit_target__cr_:
        insights.append(
            f"Deposit performance is ahead of target by "
            f"{round((b.total_deposits_cr / b.deposit_target__cr_ - 1) * 100, 1)}%."
        )
    else:
        insights.append(
            "Deposit growth is below target and needs focused mobilisation."
        )

    # Advances
    if b.advancescr >= b.advance_target:
        insights.append("Advances growth is strong and above target.")
    else:
        insights.append("Advances growth is lagging and needs acceleration.")

    # Asset Quality
    if b["npa_%"] < 3:
        insights.append("Asset quality is healthy with controlled NPAs.")
    elif b["npa_%"] < 6:
        insights.append("NPAs are moderately elevated; close monitoring required.")
    else:
        insights.append("High NPAs observed; immediate corrective action required.")

    # Profitability
    if b.profit_per_staff >= 5:
        insights.append("Profitability per staff is healthy.")
    else:
        insights.append("Profitability per staff is low, indicating efficiency gaps.")

    return insights[:4]  # keep it crisp

def generate_executive_summary(b):
    parts = []

    parts.append(
        f"Branch {b.branch_id} located in {b.city} ({b.zone} Zone) "
        f"has shown {'strong' if b.total_deposits_cr >= b.deposit_target__cr_ else 'moderate'} "
        f"business performance during the review period."
    )

    if b.total_deposits_cr >= b.deposit_target__cr_:
        parts.append(
            "Deposit mobilisation is above target, indicating healthy customer acquisition and retention."
        )
    else:
        parts.append(
            "Deposit mobilisation remains below target and requires focused efforts."
        )

    if b.advancescr >= b.advance_target:
        parts.append(
            "Advances growth is robust and supports overall balance sheet expansion."
        )
    else:
        parts.append(
            "Advances growth is lagging and needs acceleration."
        )

    if b["npa_%"] < 3:
        parts.append(
            "Asset quality remains healthy with NPAs well within acceptable limits."
        )
    elif b["npa_%"] < 6:
        parts.append(
            "Asset quality indicators show moderately elevated NPAs, requiring close monitoring."
        )
    else:
        parts.append(
            "Asset quality is under stress with high NPAs, requiring immediate corrective measures."
        )

    if b.profit_per_staff >= 5:
        parts.append(
            "Profitability indicators are satisfactory with healthy profit per staff."
        )
    else:
        parts.append(
            "Profit per staff remains below benchmark levels, indicating scope for operational efficiency improvements."
        )

    return " ".join(parts)

def score_deposits(actual, target):
    return min((actual / target) * 30, 30)

def score_advances(actual, target):
    return min((actual / target) * 25, 25)

def score_npa(npa):
    if npa <= 3:
        return 25
    elif npa <= 6:
        return 15
    else:
        return 5

def score_profitability(profit_per_staff):
    if profit_per_staff >= 5:
        return 20
    elif profit_per_staff >= 3:
        return 12
    else:
        return 5

def calculate_branch_score(b):
    score = 0
    score += score_deposits(b.total_deposits_cr, b.deposit_target__cr_)
    score += score_advances(b.advancescr, b.advance_target)
    score += score_npa(b["npa_%"])
    score += score_profitability(b.profit_per_staff)
    return round(score, 1)

def grade_branch(score):
    if score >= 80:
        return "A", "Excellent overall performance with strong fundamentals."
    elif score >= 65:
        return "B", "Good performance with minor improvement areas."
    elif score >= 50:
        return "C", "Average performance; focused corrective action required."
    else:
        return "D", "Weak performance; immediate management intervention needed."
    
def generate_risk_and_focus(b):
    risks = []
    focus = []

    # --- Profitability ---
    if b.profit_per_staff < 3:
        risks.append("Low profit per staff impacting overall efficiency.")
        focus.append("Improve staff productivity and cross-selling.")
    elif b.profit_per_staff < 5:
        focus.append("Enhance fee income and operational efficiency.")

    # --- Asset Quality ---
    if b["npa_%"] > 6:
        risks.append("High NPA posing asset quality risk.")
        focus.append("Immediate recovery actions and SMA monitoring.")
    elif b["npa_%"] > 3:
        risks.append("Moderately elevated NPA requiring close monitoring.")
        focus.append("Strengthen credit monitoring and early warning systems.")

    # --- Deposits ---
    casa_ratio = (b.total_deposits_cr * 0.5) / b.total_deposits_cr * 100  # proxy
    if casa_ratio < 40:
        risks.append("Low CASA ratio affecting cost of funds.")
        focus.append("Focused CASA mobilisation drives.")

    # --- Advances ---
    if b.advancescr < b.advance_target:
        risks.append("Advances growth below target.")
        focus.append("Push quality retail and MSME credit growth.")

    if not risks:
        risks.append("No major risk drivers identified.")
        focus.append("Sustain current performance levels.")

    return risks, focus


# ---------------- CONFIG ----------------
BRANCH_ID = sys.argv[1] if len(sys.argv) > 1 else "B1001"
OUTPUT_DIR = sys.argv[2] if len(sys.argv) > 2 else "generated"

OUTPUT_FILE = os.path.join(
    OUTPUT_DIR,
    f"Branch_Profile_{BRANCH_ID}.xlsx"
)
DATA_FILE = "../data/processed/Branch_Profile.xlsx"

# ---------------- LOAD DATA ----------------
df = pd.read_excel(DATA_FILE, sheet_name="Branch_Profile")
branch = df[df["branch_id"] == BRANCH_ID]

if branch.empty:
    raise ValueError(f"Branch {BRANCH_ID} not found")

b = branch.iloc[0]

# ---------------- EXCEL SETUP ----------------
wb = Workbook()
ws = wb.active
ws.title = "Branch Profile"

thin = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

section_fill = PatternFill("solid", fgColor="D9E1F2")
header_fill = PatternFill("solid", fgColor="BDD7EE")

def section_title(row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, size=12)
    c.alignment = Alignment(horizontal="center")
    c.fill = section_fill
    for col in range(1, 9):
        ws.cell(row=row, column=col).border = thin

def label(row, col, text):
    c = safe_cell(row, col)
    c.value = text
    c.font = Font(bold=True)
    c.border = thin

def value(row, col, val):
    c = safe_cell(row, col)
    c.value = val
    c.border = thin


def table_header(row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True)
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")
    c.border = thin

# ---------------- HEADER ----------------
ws.merge_cells("A1:H1")
ws["A1"] = "BANK OF INDIA"
ws["A1"].font = Font(size=16, bold=True)
ws["A1"].alignment = Alignment(horizontal="center")

ws.merge_cells("A2:H2")
# ---------------- INPUT + STATUS ROW ----------------
branch_score = calculate_branch_score(b)
grade, grade_remark = grade_branch(branch_score)

# ---------------- HEADER ----------------
ws.merge_cells("A1:H1")
ws["A1"] = "BANK OF INDIA"
ws["A1"].font = Font(size=16, bold=True)
ws["A1"].alignment = Alignment(horizontal="center")

ws.merge_cells("A2:H2")
ws["A2"] = f"BRANCH PROFILE AS ON : {date.today().strftime('%d-%b-%Y')}"
ws["A2"].alignment = Alignment(horizontal="center")

# --- Branch Code Input (Editable) ---
ws["A3"] = "Branch Code:"
ws["A3"].font = Font(bold=True)

ws["B3"] = BRANCH_ID
ws["B3"].border = thin

# --- Grade & Score (Read-only, right side) ---
ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=8)
status = ws.cell(
    row=3,
    column=4,
    value=f"Overall Grade: {grade}   |   Score: {branch_score}/100"
)

status.font = Font(bold=True)
status.alignment = Alignment(horizontal="center", vertical="center")
status.border = thin

status.fill = (
    green if grade == "A" else
    amber if grade == "B" else
    PatternFill("solid", fgColor="F4B084") if grade == "C" else
    red
)

status.border = thin

# ---------------- STATUS STRIP ----------------
branch_score = calculate_branch_score(b)
grade, grade_remark = grade_branch(branch_score)


status = ws.cell(
    row=3,
    column=1,
    value=f"Branch Code: {BRANCH_ID}        Overall Grade: {grade}  |  Score: {branch_score}/100"
)

status.font = Font(bold=True)
status.alignment = Alignment(horizontal="center", vertical="center")

status.fill = (
    green if grade == "A" else
    amber if grade == "B" else
    PatternFill("solid", fgColor="F4B084") if grade == "C" else
    red
)

status.border = thin

ws["A2"] = f"BRANCH PROFILE AS ON : {date.today().strftime('%d-%b-%Y')}"
ws["A2"].alignment = Alignment(horizontal="center")


# ---------------- KEY TAKEAWAYS ----------------
section_title(5, "KEY TAKEAWAYS")

takeaways = generate_key_takeaways(b)

kt_row = 7
END_KT_ROW = kt_row
for point in takeaways:
    ws.merge_cells(start_row=kt_row, start_column=1, end_row=kt_row, end_column=8)
    cell = ws.cell(row=kt_row, column=1, value=f"• {point}")
    cell.border = thin
    cell.alignment = Alignment(wrap_text=True)
    kt_row += 1

# ---------------- EXECUTIVE SUMMARY ----------------
EXEC_SUMMARY_ROW = kt_row + 1

section_title(EXEC_SUMMARY_ROW, "EXECUTIVE SUMMARY")

summary_text = generate_executive_summary(b)

ws.merge_cells(
    start_row=EXEC_SUMMARY_ROW + 2,
    start_column=1,
    end_row=EXEC_SUMMARY_ROW + 5,
    end_column=8
)

cell = ws.cell(row=EXEC_SUMMARY_ROW + 2, column=1, value=summary_text)
cell.alignment = Alignment(wrap_text=True, vertical="top")
cell.border = thin

def safe_cell(row, col):
    cell = ws.cell(row=row, column=col)
    if cell.coordinate in ws.merged_cells:
        raise ValueError(f"Attempting to write into merged cell {cell.coordinate}")
    return cell

# ---------------- BRANCH DETAILS ----------------
BRANCH_START_ROW = EXEC_SUMMARY_ROW + 6

section_title(BRANCH_START_ROW, "BRANCH DETAILS")

label(BRANCH_START_ROW + 2, 1, "Branch ID")
value(BRANCH_START_ROW + 2, 2, b.branch_id)

label(BRANCH_START_ROW + 2, 3, "Branch Name")
value(BRANCH_START_ROW + 2, 4, b.branch_name)

label(BRANCH_START_ROW + 3, 1, "Zone")
value(BRANCH_START_ROW + 3, 2, b.zone)

label(BRANCH_START_ROW + 3, 3, "City")
value(BRANCH_START_ROW + 3, 4, b.city)

label(BRANCH_START_ROW + 4, 1, "Risk Category")
risk_cell = ws.cell(row=BRANCH_START_ROW + 4, column=2, value=b.risk_flag)
risk_cell.border = thin
risk_cell.fill = green if b.risk_flag == "Healthy" else amber if b.risk_flag == "Watch" else red

label(BRANCH_START_ROW + 4, 3, "NPA %")
value(BRANCH_START_ROW + 4, 4, round(b["npa_%"], 2))

    
# ---------------- KPI SCORECARD & RATING ----------------
KPI_ROW = BRANCH_START_ROW + 7

section_title(KPI_ROW, "BRANCH KPI SCORECARD & RATING")

scorecard_row = KPI_ROW + 2

headers_kpi = ["KPI", "Actual", "Target / Benchmark", "Status"]
for i, h in enumerate(headers_kpi):
    table_header(scorecard_row, 1 + i, h)

scorecard_row += 1

def kpi_status(actual, target, reverse=False):
    if reverse:  # for NPA
        if actual <= target:
            return "Good", green
        elif actual <= target * 2:
            return "Moderate", amber
        else:
            return "High Risk", red
    else:
        if actual >= target:
            return "Ahead", green
        elif actual >= target * 0.9:
            return "Slight Lag", amber
        else:
            return "Behind", red

kpis = [
    ("Deposits (₹ Cr)", b.total_deposits_cr, b.deposit_target__cr_, False),
    ("Advances (₹ Cr)", b.advancescr, b.advance_target, False),
    ("NPA %", round(b["npa_%"], 2), 3, True),
    ("Profit / Staff", round(b.profit_per_staff, 2), 5, False),
]

for name, actual, target, reverse in kpis:
    ws.cell(row=scorecard_row, column=1, value=name).border = thin
    ws.cell(row=scorecard_row, column=2, value=actual).border = thin
    ws.cell(row=scorecard_row, column=3, value=target).border = thin

    status, fill = kpi_status(actual, target, reverse)
    c = ws.cell(row=scorecard_row, column=4, value=status)
    c.border = thin
    c.fill = fill

    scorecard_row += 1


# ---------------- KEY RISK DRIVERS & FOCUS AREAS ----------------
risk_row = scorecard_row + 2

section_title(risk_row, "KEY RISK DRIVERS & PRIORITY FOCUS AREAS")

risks, focus_areas = generate_risk_and_focus(b)

# --- Risk Drivers ---
ws.merge_cells(start_row=risk_row + 2, start_column=1, end_row=risk_row + 2, end_column=8)
r = ws.cell(row=risk_row + 2, column=1, value="🔴 KEY RISK DRIVERS")
r.font = Font(bold=True)
r.border = thin

row_ptr = risk_row + 3
for risk in risks:
    ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=8)
    c = ws.cell(row=row_ptr, column=1, value=f"• {risk}")
    c.border = thin
    row_ptr += 1

# --- Focus Areas ---
ws.merge_cells(start_row=row_ptr + 1, start_column=1, end_row=row_ptr + 1, end_column=8)
f = ws.cell(row=row_ptr + 1, column=1, value="🟢 PRIORITY FOCUS AREAS (Next 90 Days)")
f.font = Font(bold=True)
f.border = thin

row_ptr += 2
for area in focus_areas:
    ws.merge_cells(start_row=row_ptr, start_column=1, end_row=row_ptr, end_column=8)
    c = ws.cell(row=row_ptr, column=1, value=f"• {area}")
    c.border = thin
    row_ptr += 1


# ---- KPI RESULT ROW ----
ws.merge_cells(start_row=scorecard_row, start_column=1, end_row=scorecard_row, end_column=2)
ws.merge_cells(start_row=scorecard_row, start_column=3, end_row=scorecard_row, end_column=4)

score_cell = ws.cell(
    row=scorecard_row,
    column=1,
    value=f"OVERALL KPI SCORE : {branch_score}/100"
)
score_cell.font = Font(bold=True)
score_cell.border = thin

grade_cell = ws.cell(
    row=scorecard_row,
    column=3,
    value=f"GRADE : {grade} – {grade_remark}"
)
grade_cell.alignment = Alignment(wrap_text=True)
grade_cell.border = thin

score_cell.fill = grade_cell.fill = (
    green if grade == "A" else
    amber if grade == "B" else
    PatternFill("solid", fgColor="F4B084") if grade == "C" else
    red
)


# ---------------- DEPOSITS (TABULAR MOCK) ----------------
row = row_ptr + 1
section_title(row, "DEPOSITS POSITION (₹ Crores)")
row += 2

headers = ["Particulars", "Actual", "Target", "Achievement %", "GAP"]
for i, h in enumerate(headers):
    table_header(row, 1 + i, h)

row += 1

# --- LOGICAL BREAKUP (MOCK BUT CONSISTENT) ---
savings = round(b.total_deposits_cr * 0.35, 2)
current = round(b.total_deposits_cr * 0.15, 2)
casa = round(savings + current, 2)
td = round(b.total_deposits_cr - casa, 2)

deposit_rows = [
    ("Savings Deposits", savings),
    ("Current Deposits", current),
    ("CASA Deposits (Savings + Current)", casa),
    ("Term Deposits", td),
    ("TOTAL DEPOSITS", b.total_deposits_cr),
]

for name, actual in deposit_rows:
    target = round(b.deposit_target__cr_, 2)
    achievement = round((actual / target) * 100, 2)
    gap = round(target - actual, 2)

    ws.cell(row=row, column=1, value=name).border = thin
    ws.cell(row=row, column=2, value=round(actual, 2)).border = thin
    ws.cell(row=row, column=3, value=target).border = thin

    ach_cell = ws.cell(row=row, column=4, value=achievement)
    gap_cell = ws.cell(row=row, column=5, value=gap)

    ach_cell.border = thin
    gap_cell.border = thin

    colour_achievement(ach_cell, achievement)
    colour_gap(gap_cell, gap)

    row += 1


# ---------------- ADVANCES (TABULAR) ----------------
section_title(row + 1, "ADVANCES POSITION (₹ Crores)")
row += 3

for i, h in enumerate(headers):
    table_header(row, 1 + i, h)

row += 1
ws.cell(row=row, column=1, value="TOTAL ADVANCES").border = thin
ws.cell(row=row, column=2, value=round(b.advancescr, 2)).border = thin
ws.cell(row=row, column=3, value=round(b.advance_target, 2)).border = thin
ach = round(b.advance_ach_pct, 2)
adv_gap = round(b.advance_target - b.advancescr, 2)

ach_cell = ws.cell(row=row, column=4, value=ach)
gap_cell = ws.cell(row=row, column=5, value=adv_gap)

ach_cell.border = thin
gap_cell.border = thin

colour_achievement(ach_cell, ach)
colour_gap(gap_cell, adv_gap)


# ---------------- STAFF & PROFIT ----------------
# ---------------- ASSET QUALITY & PERFORMANCE ----------------
section_title(row + 8, "ASSET QUALITY & PERFORMANCE")

aq_row = row + 10

label(aq_row, 1, "NPA Level (%)")
value(aq_row, 2, round(b["npa_%"], 2))

label(aq_row, 3, "Risk Category")
value(aq_row, 4, b.risk_flag)

# ---- Interpretations ----
if b["npa_%"] < 3:
    npa_comment = "NPA level is within acceptable limits."
elif b["npa_%"] < 6:
    npa_comment = "NPA slightly elevated. Close monitoring required."
else:
    npa_comment = "High NPA. Immediate corrective action required."

ws.merge_cells(start_row=aq_row + 1, start_column=1, end_row=aq_row + 1, end_column=8)
c = ws.cell(row=aq_row + 1, column=1, value=npa_comment)
c.border = thin
c.alignment = Alignment(wrap_text=True)

# ---------------- PERFORMANCE FLAGS ----------------
section_title(aq_row + 3, "PERFORMANCE FLAGS")

pf_row = aq_row + 5

def flag(actual, target):
    if actual >= target:
        return "Ahead of Target"
    elif actual >= target * 0.9:
        return "Slightly Behind"
    else:
        return "Needs Immediate Attention"

label(pf_row, 1, "Deposits Performance")
cell = ws.cell(row=pf_row, column=2, value=flag(b.total_deposits_cr, b.deposit_target__cr_))
cell.border = thin
cell.fill = green if "Ahead" in cell.value else amber if "Slightly" in cell.value else red


label(pf_row + 1, 1, "Advances Performance")
cell = ws.cell(row=pf_row, column=2, value=flag(b.advancescr, b.advance_target))
cell.border = thin
cell.fill = green if "Ahead" in cell.value else amber if "Slightly" in cell.value else red


label(pf_row + 2, 1, "Profitability Status")
cell = ws.cell(row=pf_row, column=2, value=flag(b.profit_per_staff, 5))
cell.border = thin
cell.fill = green if "Ahead" in cell.value else amber if "Slightly" in cell.value else red



# ---------------- OFFICER REMARKS ----------------
section_title(pf_row + 4, "OFFICER REMARKS")

remarks = []

if b.total_deposits_cr < b.deposit_target__cr_:
    remarks.append("Deposit growth below target.")
else:
    remarks.append("Deposit performance satisfactory.")

if b.advancescr >= b.advance_target:
    remarks.append("Advances growth strong.")

if b["npa_%"] > 5:
    remarks.append("Asset quality needs close monitoring.")

if not remarks:
    remarks.append("Overall performance satisfactory.")

ws.merge_cells(start_row=pf_row + 6, start_column=1, end_row=pf_row + 8, end_column=8)
c = ws.cell(row=pf_row + 6, column=1, value=" ".join(remarks))
c.alignment = Alignment(wrap_text=True)
c.border = thin

section_title(row + 3, "STAFF & PROFITABILITY")

label(row + 5, 1, "Staff Strength")
value(row + 5, 2, int(b.staff_strength))

label(row + 5, 3, "Total Profit (₹ Cr)")
value(row + 5, 4, round(b.profit_cr, 2))

label(row + 6, 1, "Profit per Staff")
value(row + 6, 2, round(b.profit_per_staff, 2))

# ---------------- FORMAT ----------------
for col in range(1, 9):
    ws.column_dimensions[chr(64 + col)].width = 20

os.makedirs(OUTPUT_DIR, exist_ok=True)

wb.save(OUTPUT_FILE)
print(f"\n✅ STEP-4B COMPLETE: {OUTPUT_FILE}")
